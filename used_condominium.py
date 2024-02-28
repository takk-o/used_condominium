from pathlib import Path
import openpyxl as px
import requests
from bs4 import BeautifulSoup

import ssl
context = ssl.SSLContext(ssl.PROTOCOL_TLSv1)

def read_workbook(rd_fld, rd_fle):
    rd_fld_pth = Path(rd_fld)
    rd_fle_pth = rd_fld_pth.joinpath(rd_fle)        
    wb = px.load_workbook(rd_fle_pth)
    return wb

url = 'https://www.mansion-review.jp/search/'
soup = BeautifulSoup(requests.get(url, verify=False).content, 'html.parser')
div = soup.select('div.area_search_tizu_container')
prefs = div[0].select('a')
dic_pref = {}
for pref in prefs:
    dic_pref[pref.text] = pref.get('href')

wb = read_workbook('.', 'used_condo.xlsx')
ws = wb['used_condo']
search_pref = ws["C2"].value

url = dic_pref[search_pref]
soup = BeautifulSoup(requests.get(url, verify=False).content, 'html.parser')
tab1 = soup.select('#tab1')
elements = tab1[0].select('.cb_list_B_inner')
cities = [element.select_one('a') for element in elements]
dic_city = {}
for city in cities:
    dic_city[city.text] = city.get('href')

search_city = ws["E2"].value
for city, city_url in dic_city.items():
    if search_city in city:
        break
soup = BeautifulSoup(requests.get(city_url, verify=False).content, 'html.parser')
result_list = soup.select('ul.searchResultList li.property-detail-list-item')
row = 6
for element1 in result_list:
    ws.cell(row=row, column=2, value=element1.select_one('a').text)
    element2 = element1.select('table.property-detail-content_main td')
    ws.cell(row=row, column=3, value=element2[0].text)
    ws.cell(row=row, column=4, value=' '.join([element3.text for element3 in element2[1].select('span')]))
    ws.cell(row=row, column=5, value=element2[2].text)
    ws.cell(row=row, column=6, value=element2[3].text.strip())
    ws.cell(row=row, column=7, value=element2[4].text)
    row += 1

wb.save('used_condo.xlsx')
wb.close()